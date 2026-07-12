"""XLSX export helpers for the final planning workbook."""

from __future__ import annotations

from io import BytesIO
from math import ceil
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.allocation import RAPlan


WEEKDAY_ABBREVIATIONS = {
    "Dilluns": "dl.",
    "Dimarts": "dt.",
    "Dimecres": "dc.",
    "Dijous": "dj.",
    "Divendres": "dv.",
}

RA_HEADER_FILLS = [
    "E7EFD8",
    "D8EBF2",
    "F7E3D1",
    "E2D8F0",
    "E8E0BC",
    "F6D9E4",
    "D7EEE7",
    "F6E9C9",
    "DDE3F5",
    "F5DCCF",
    "DCECD3",
    "EFE0C8",
]
ROW_START_FILL = PatternFill(fill_type="solid", fgColor="E9E9E9")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
CALENDAR_METADATA_FIELDS = (
    "Cicle formatiu",
    "Grup",
    "Codi del mòdul",
    "Nom del mòdul",
)
CALENDAR_HEADER_ROW = 6
CALENDAR_DATA_START_ROW = 7
SUMMARY_LABELS = (
    "Hores previstes:",
    "Hores reals:",
    "Acompliment:",
)
SHEET_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
COMMENT_COLUMN_WIDTH_MULTIPLIER = 3
RA_COLUMN_WIDTH = 14
INSTRUCTION_BANNER = (
    "MODIFIQUEU LES HORES D'UNA DATA CONCRETA PER RECALCULAR AUTOMÀTICAMENT LA RÀTIO D’ACOMPLIMENT"
)


def _auto_width_for_column_values(values: list[object]) -> float:
    """Return a simple width based on the visible text length."""

    visible_lengths = [len(str(value)) for value in values if value is not None]
    longest = max(visible_lengths, default=0)
    return max(longest + 2, 6)


def _instruction_banner_chars_per_line(ra_count: int) -> int:
    """Estimate how many visible characters fit per banner line."""

    merged_width = (ra_count * RA_COLUMN_WIDTH) + (RA_COLUMN_WIDTH * COMMENT_COLUMN_WIDTH_MULTIPLIER)
    return max(int(merged_width * 0.95), 24)


def _instruction_banner_text(ra_count: int) -> str:
    """Insert a balanced line break before Excel wraps the banner awkwardly."""

    chars_per_line = _instruction_banner_chars_per_line(ra_count)
    if "\n" in INSTRUCTION_BANNER or len(INSTRUCTION_BANNER) <= chars_per_line:
        return INSTRUCTION_BANNER

    words = INSTRUCTION_BANNER.split()
    if len(words) < 2:
        return INSTRUCTION_BANNER

    best_left = INSTRUCTION_BANNER
    best_right = ""
    best_score = None
    for index in range(1, len(words)):
        left = " ".join(words[:index])
        right = " ".join(words[index:])
        score = (max(len(left), len(right)), abs(len(left) - len(right)))
        if best_score is None or score < best_score:
            best_left = left
            best_right = right
            best_score = score

    return f"{best_left}\n{best_right}"


def _instruction_banner_height(ra_count: int) -> float:
    """Estimate a readable row height for the wrapped banner text."""

    chars_per_line = _instruction_banner_chars_per_line(ra_count)
    line_count = sum(
        max(1, ceil(len(line) / chars_per_line))
        for line in _instruction_banner_text(ra_count).splitlines()
    )
    return max(24, line_count * 18)


def _active_ra_keys(row: dict[str, object], ras: list[RAPlan]) -> tuple[str, ...]:
    """Return the ordered RA keys with assigned hours in one calendar row."""

    ra_hours = row["ra_hours"]
    return tuple(ra.key for ra in ras if ra_hours.get(ra.key, 0))


def _calendar_metadata_value(summary: dict[str, object], key: str) -> object:
    """Return a blank cell for missing optional metadata instead of a placeholder."""

    value = summary.get(key)
    if value in {None, "-"}:
        return None
    return value


def _apply_ra_fill(cell, index: int) -> None:
    """Apply the repeating RA palette fill to one cell."""

    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=RA_HEADER_FILLS[index % len(RA_HEADER_FILLS)],
    )


def _suppress_adjacent_formula_warnings(workbook_bytes: BytesIO, ignored_ranges: list[str]) -> BytesIO:
    """Mark selected formula ranges as intentionally ignored for Excel warnings."""

    if not ignored_ranges:
        workbook_bytes.seek(0)
        return workbook_bytes

    workbook_bytes.seek(0)
    source = BytesIO(workbook_bytes.getvalue())
    patched = BytesIO()
    ET.register_namespace("", SHEET_MAIN_NS)
    worksheet_path = "xl/worksheets/sheet1.xml"

    with ZipFile(source, "r") as source_zip, ZipFile(patched, "w", ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename == worksheet_path:
                root = ET.fromstring(data)
                ignored_errors = root.find(f"{{{SHEET_MAIN_NS}}}ignoredErrors")
                if ignored_errors is None:
                    ignored_errors = ET.Element(f"{{{SHEET_MAIN_NS}}}ignoredErrors")
                    root.append(ignored_errors)
                for cell_range in ignored_ranges:
                    ET.SubElement(
                        ignored_errors,
                        f"{{{SHEET_MAIN_NS}}}ignoredError",
                        {"sqref": cell_range, "formulaRange": "1"},
                    )
                data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
            target_zip.writestr(item, data)

    patched.seek(0)
    return patched


def build_workbook(
    calendar_rows: list[dict[str, object]],
    ras: list[RAPlan],
    summary: dict[str, object],
    *,
    include_week_numbers: bool = False,
) -> BytesIO:
    """Create the two-sheet workbook returned to the teacher.

    Zero values in RA cells are written as empty cells so the resulting calendar
    is easier to scan visually. When `include_week_numbers` is set, each row's
    `week_number` key (if present) is rendered as a leading "Set." column;
    callers should only pass it when the admin's official week numbering
    actually applies to the planned date range.
    """

    workbook = Workbook()
    calendar_sheet = workbook.active
    calendar_sheet.title = "Calendari"
    column_offset = 1 if include_week_numbers else 0
    week_column = 1
    weekday_column = 1 + column_offset
    date_column = 2 + column_offset
    hours_column = 3 + column_offset
    first_ra_column = 4 + column_offset
    comment_column_index = first_ra_column + len(ras)
    has_optional_ra_names = any((ra.name or "").strip() != ra.key for ra in ras)
    header_main_row = CALENDAR_HEADER_ROW
    header_sub_row = CALENDAR_HEADER_ROW + 1 if has_optional_ra_names else None
    data_start_row = CALENDAR_DATA_START_ROW + (1 if has_optional_ra_names else 0)

    calendar_sheet.merge_cells(start_row=1, start_column=first_ra_column, end_row=1, end_column=comment_column_index)
    instruction_cell = calendar_sheet.cell(row=1, column=first_ra_column)
    instruction_cell.value = _instruction_banner_text(len(ras))
    instruction_cell.font = Font(bold=True, color="FFFFFF")
    instruction_cell.fill = PatternFill(fill_type="solid", fgColor="000000")
    instruction_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    calendar_sheet.row_dimensions[1].height = _instruction_banner_height(len(ras))

    for row_index, field_name in enumerate(CALENDAR_METADATA_FIELDS, start=2):
        calendar_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=hours_column)
        label_cell = calendar_sheet.cell(row=row_index, column=1)
        label_cell.value = f"{field_name}:"
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell = calendar_sheet.cell(row=row_index, column=first_ra_column)
        value_cell.value = _calendar_metadata_value(summary, field_name)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    if has_optional_ra_names:
        calendar_sheet.merge_cells(
            start_row=header_main_row, start_column=weekday_column, end_row=header_sub_row, end_column=date_column
        )
        calendar_sheet.merge_cells(
            start_row=header_main_row, start_column=hours_column, end_row=header_sub_row, end_column=hours_column
        )
        if include_week_numbers:
            calendar_sheet.merge_cells(
                start_row=header_main_row, start_column=week_column, end_row=header_sub_row, end_column=week_column
            )
        calendar_sheet.merge_cells(
            start_row=header_main_row,
            start_column=comment_column_index,
            end_row=header_sub_row,
            end_column=comment_column_index,
        )
    else:
        calendar_sheet.merge_cells(
            start_row=header_main_row, start_column=weekday_column, end_row=header_main_row, end_column=date_column
        )

    if include_week_numbers:
        week_header_cell = calendar_sheet.cell(row=header_main_row, column=week_column)
        week_header_cell.value = "Set."
        week_header_cell.font = Font(bold=True)
        week_header_cell.alignment = Alignment(horizontal="center", vertical="center")

    data_header_cell = calendar_sheet.cell(row=header_main_row, column=weekday_column)
    data_header_cell.value = "Data"
    data_header_cell.alignment = Alignment(horizontal="left", vertical="center")
    data_header_cell.font = Font(bold=True)

    hours_header_cell = calendar_sheet.cell(row=header_main_row, column=hours_column)
    hours_header_cell.value = "Hores"
    hours_header_cell.alignment = Alignment(horizontal="center", vertical="center")
    hours_header_cell.font = Font(bold=True)

    comment_header_cell = calendar_sheet.cell(row=header_main_row, column=comment_column_index)
    comment_header_cell.value = "Comentaris"
    comment_header_cell.alignment = Alignment(horizontal="center", vertical="center")
    comment_header_cell.font = Font(bold=True)

    for index, ra in enumerate(ras):
        column_index = first_ra_column + index
        key_cell = calendar_sheet.cell(row=header_main_row, column=column_index)
        key_cell.value = ra.key
        key_cell.font = Font(bold=True)
        key_cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_ra_fill(key_cell, index)

        if has_optional_ra_names:
            optional_name = (ra.name or "").strip()
            if optional_name != ra.key:
                name_cell = calendar_sheet.cell(row=header_sub_row, column=column_index)
                name_cell.value = optional_name
                name_cell.font = Font(bold=True)
                name_cell.alignment = Alignment(horizontal="center", vertical="center")
                _apply_ra_fill(name_cell, index)
            else:
                calendar_sheet.merge_cells(
                    start_row=header_main_row,
                    start_column=column_index,
                    end_row=header_sub_row,
                    end_column=column_index,
                )

    calendar_sheet.freeze_panes = f"{get_column_letter(first_ra_column)}{data_start_row}"

    ra_column_by_key = {ra.key: first_ra_column + index for index, ra in enumerate(ras)}
    started_ra_keys: set[str] = set()
    for row in calendar_rows:
        excel_row = []
        if include_week_numbers:
            excel_row.append(row.get("week_number"))
        excel_row.extend([WEEKDAY_ABBREVIATIONS.get(row["weekday"], row["weekday"]), row["date"], row["total_hours"]])
        ra_hours = row["ra_hours"]
        excel_row.extend(ra_hours[ra.key] or None for ra in ras)
        excel_row.append(None)
        calendar_sheet.append(excel_row)
        active_keys = _active_ra_keys(row, ras)
        newly_started_keys = [key for key in active_keys if key not in started_ra_keys]
        if newly_started_keys:
            current_row = calendar_sheet.max_row
            for cell in calendar_sheet[current_row]:
                cell.fill = ROW_START_FILL
            for key in newly_started_keys:
                column_index = ra_column_by_key[key]
                _apply_ra_fill(calendar_sheet.cell(row=current_row, column=column_index), column_index - first_ra_column)
        started_ra_keys.update(active_keys)

    summary_start_row = calendar_sheet.max_row + 1
    for row_offset, label in enumerate(SUMMARY_LABELS):
        row_number = summary_start_row + row_offset
        calendar_sheet.cell(row=row_number, column=1).value = label
        calendar_sheet.cell(row=row_number, column=1).font = Font(bold=True)
        calendar_sheet.cell(row=row_number, column=1).alignment = Alignment(horizontal="left", vertical="center")
        calendar_sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=hours_column)

    data_end_row = summary_start_row - 1
    expected_row = summary_start_row
    actual_row = summary_start_row + 1
    completion_row = summary_start_row + 2
    for index, ra in enumerate(ras):
        column_index = first_ra_column + index
        column_letter = get_column_letter(column_index)
        expected_cell = calendar_sheet.cell(row=expected_row, column=column_index)
        expected_cell.value = ra.hours
        expected_cell.font = Font(bold=True)
        expected_cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_ra_fill(expected_cell, index)

        actual_cell = calendar_sheet.cell(row=actual_row, column=column_index)
        actual_cell.value = f"=SUM({column_letter}{data_start_row}:{column_letter}{data_end_row})"
        actual_cell.alignment = Alignment(horizontal="center", vertical="center")

        completion_cell = calendar_sheet.cell(row=completion_row, column=column_index)
        completion_cell.value = f"=IFERROR({column_letter}{actual_row}/{column_letter}{expected_row},\"\")"
        completion_cell.number_format = "0%"
        completion_cell.alignment = Alignment(horizontal="center", vertical="center")

    if ras:
        first_ra_column_letter = get_column_letter(first_ra_column)
        last_ra_column_letter = get_column_letter(first_ra_column - 1 + len(ras))
        comment_column_letter = get_column_letter(comment_column_index)

        comment_expected_cell = calendar_sheet.cell(row=expected_row, column=comment_column_index)
        comment_expected_cell.value = (
            f"=SUM({first_ra_column_letter}{expected_row}:{last_ra_column_letter}{expected_row})"
        )
        comment_expected_cell.font = Font(bold=True)
        comment_expected_cell.alignment = Alignment(horizontal="center", vertical="center")
        comment_expected_cell.number_format = "0.00"

        comment_actual_cell = calendar_sheet.cell(row=actual_row, column=comment_column_index)
        comment_actual_cell.value = (
            f"=SUM({first_ra_column_letter}{actual_row}:{last_ra_column_letter}{actual_row})"
        )
        comment_actual_cell.alignment = Alignment(horizontal="center", vertical="center")
        comment_actual_cell.number_format = "0.00"

        comment_completion_cell = calendar_sheet.cell(row=completion_row, column=comment_column_index)
        comment_completion_cell.value = (
            f"=IFERROR({comment_column_letter}{actual_row}/{comment_column_letter}{expected_row},\"\")"
        )
        comment_completion_cell.number_format = "0%"
        comment_completion_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in calendar_sheet.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=date_column, max_col=date_column):
        row[0].number_format = "DD/MM/YYYY"

    weekday_letter = get_column_letter(weekday_column)
    date_letter = get_column_letter(date_column)
    hours_letter = get_column_letter(hours_column)
    weekday_values = [calendar_sheet[f"{weekday_letter}{row_index}"].value for row_index in range(data_start_row, data_end_row + 1)]
    date_values = [calendar_sheet[f"{date_letter}{row_index}"].value.strftime("%d/%m/%Y") for row_index in range(data_start_row, data_end_row + 1)]
    hour_values = [calendar_sheet[f"{hours_letter}{row_index}"].value for row_index in range(header_main_row, data_end_row + 1)]
    calendar_sheet.column_dimensions[weekday_letter].width = _auto_width_for_column_values(weekday_values)
    calendar_sheet.column_dimensions[date_letter].width = _auto_width_for_column_values(date_values)
    calendar_sheet.column_dimensions[hours_letter].width = _auto_width_for_column_values(hour_values)
    if include_week_numbers:
        week_letter = get_column_letter(week_column)
        week_values = [calendar_sheet[f"{week_letter}{row_index}"].value for row_index in range(header_main_row, data_end_row + 1)]
        calendar_sheet.column_dimensions[week_letter].width = _auto_width_for_column_values(week_values)
    for row in calendar_sheet.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=1, max_col=date_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in calendar_sheet.iter_rows(min_row=header_main_row, max_row=completion_row, min_col=hours_column, max_col=hours_column):
        row[0].alignment = Alignment(horizontal="center", vertical="center")
    for index in range(len(ras)):
        column_letter = get_column_letter(first_ra_column + index)
        calendar_sheet.column_dimensions[column_letter].width = RA_COLUMN_WIDTH
        for row in calendar_sheet.iter_rows(
            min_row=header_main_row, max_row=completion_row, min_col=first_ra_column + index, max_col=first_ra_column + index
        ):
            row[0].alignment = Alignment(horizontal="center", vertical="center")
            if data_start_row <= row[0].row < completion_row and row[0].value is not None:
                row[0].number_format = "0.00"

    comment_column_letter = get_column_letter(comment_column_index)
    calendar_sheet.column_dimensions[comment_column_letter].width = RA_COLUMN_WIDTH * COMMENT_COLUMN_WIDTH_MULTIPLIER
    for row in calendar_sheet.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=comment_column_index, max_col=comment_column_index):
        row[0].alignment = Alignment(horizontal="left", vertical="center")

    for row in calendar_sheet.iter_rows(
        min_row=header_main_row,
        max_row=data_end_row,
        min_col=1,
        max_col=comment_column_index,
    ):
        for cell in row:
            cell.border = THIN_BORDER

    if has_optional_ra_names:
        for index, ra in enumerate(ras):
            optional_name = (ra.name or "").strip()
            if optional_name == ra.key:
                continue
            column_index = first_ra_column + index
            key_cell = calendar_sheet.cell(row=header_main_row, column=column_index)
            name_cell = calendar_sheet.cell(row=header_sub_row, column=column_index)
            key_cell.border = Border(
                left=THIN_BORDER.left,
                right=THIN_BORDER.right,
                top=THIN_BORDER.top,
                bottom=Side(style=None),
            )
            name_cell.border = Border(
                left=THIN_BORDER.left,
                right=THIN_BORDER.right,
                top=Side(style=None),
                bottom=THIN_BORDER.bottom,
            )

    if ras:
        for row in calendar_sheet.iter_rows(
            min_row=expected_row,
            max_row=completion_row,
            min_col=first_ra_column,
            max_col=first_ra_column - 1 + len(ras),
        ):
            for cell in row:
                cell.border = THIN_BORDER

    summary_sheet = workbook.create_sheet("Resum")
    summary_sheet.freeze_panes = "A2"
    summary_sheet.append(["Camp", "Valor"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    for key, value in summary.items():
        summary_sheet.append([key, value])

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 40

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    if ras:
        start_column = get_column_letter(first_ra_column)
        ignored_ranges = [
            f"{start_column}{actual_row}:{comment_column_letter}{actual_row}",
            f"{start_column}{completion_row}:{comment_column_letter}{completion_row}",
        ]
        output = _suppress_adjacent_formula_warnings(output, ignored_ranges)

    return output
