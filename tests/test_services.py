"""Focused tests for the core business rules.

These tests avoid the full web stack and instead verify the calculations that
most directly affect the teacher's exported planning file.
"""

from datetime import date, timedelta
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.date_utils import format_display_date, format_month_label, parse_date_input
from app.services.academic_weeks import (
    is_vacation_week,
    iter_week_starts,
    suggest_week_numbers,
    week_has_teaching_potential,
    week_number_for_date,
    week_start,
)
from app.services.allocation import BlockPlan, RAPlan, allocate_ra_hours, allocate_ra_hours_by_blocks, validate_ra_distribution
from app.services.calendar import build_schedule, expand_excluded_periods, total_available_hours
from app.services.export import build_workbook
from app.routes.teacher import _build_block_plans
from app.services.hours import format_minutes_label, parse_hour_minute_pair


def test_available_hours_calculation_uses_real_weekdays():
    """Only teaching weekdays after the presentation day should contribute hours."""

    schedule = build_schedule(
        date(2026, 9, 1),
        date(2026, 9, 7),
        {0: 2, 1: 3, 2: 4, 3: 1, 4: 5},
        set(),
    )
    assert total_available_hours(schedule) == 12


def test_first_planning_day_is_always_excluded_from_schedule():
    """The first planning date should act as presentation day without classes."""

    schedule = build_schedule(
        date(2026, 9, 7),
        date(2026, 9, 9),
        {0: 2, 1: 2, 2: 2, 3: 0, 4: 0},
        set(),
    )
    assert [item.date.isoformat() for item in schedule] == ["2026-09-08", "2026-09-09"]


def test_holiday_ranges_are_excluded_without_duplicates():
    """Overlapping excluded periods should not subtract the same day twice."""

    excluded = expand_excluded_periods(
        [
            (date(2026, 9, 2), date(2026, 9, 4)),
            (date(2026, 9, 4), date(2026, 9, 5)),
        ]
    )
    schedule = build_schedule(date(2026, 8, 31), date(2026, 9, 7), {0: 2, 1: 2, 2: 2, 3: 2, 4: 2}, excluded)
    assert [item.date.isoformat() for item in schedule] == ["2026-09-01", "2026-09-07"]


def test_sequential_ra_allocation_can_split_same_day():
    """A day can contain the end of one RA and the start of the next."""

    schedule = build_schedule(date(2026, 8, 31), date(2026, 9, 2), {0: 4, 1: 4, 2: 4}, set())
    rows = allocate_ra_hours(schedule, [RAPlan("RA1", "RA1", 5), RAPlan("RA2", "RA2", 3)])
    assert rows[0]["ra_hours"] == {"RA1": 4, "RA2": 0}
    assert rows[1]["ra_hours"] == {"RA1": 1, "RA2": 3}


def test_distribution_validation_requires_exact_match():
    """Export must fail if the teacher leaves hours unassigned."""

    with pytest.raises(ValueError):
        validate_ra_distribution(10, [RAPlan("RA1", "RA1", 7), RAPlan("RA2", "RA2", 2)])


def test_distribution_validation_rejects_zero_hour_ras():
    """Every RA must keep at least one assigned hour before export."""

    with pytest.raises(ValueError, match="temps assignat"):
        validate_ra_distribution(10, [RAPlan("RA1", "RA1", 10), RAPlan("RA2", "RA2", 0)])


def test_parallel_blocks_preserve_their_weekly_split_when_both_need_all_hours():
    """Parallel blocks should stay in their own lanes when neither finishes early."""

    schedule = build_schedule(
        date(2026, 9, 7),
        date(2026, 9, 21),
        {0: 2, 2: 2},
        set(),
    )
    rows = allocate_ra_hours_by_blocks(
        schedule,
        [
            RAPlan("RA1", "RA1", 2, block_key="block_1"),
            RAPlan("RA2", "RA2", 6, block_key="block_2"),
        ],
        [
            BlockPlan("block_1", "Bloc 1", {0: 60, 1: 0, 2: 0, 3: 0, 4: 0}),
            BlockPlan("block_2", "Bloc 2", {0: 60, 1: 0, 2: 120, 3: 0, 4: 0}),
        ],
    )
    assert rows[0]["ra_hours"] == {"RA1": 0, "RA2": 2}
    assert rows[1]["ra_hours"] == {"RA1": 1, "RA2": 1}
    assert rows[2]["ra_hours"] == {"RA1": 0, "RA2": 2}
    assert rows[3]["ra_hours"] == {"RA1": 1, "RA2": 1}


def test_parallel_blocks_automatically_transfer_unused_hours():
    """Unused hours from one block should be absorbed by the other block automatically."""

    schedule = build_schedule(
        date(2026, 9, 7),
        date(2026, 9, 21),
        {0: 2, 2: 2},
        set(),
    )
    rows = allocate_ra_hours_by_blocks(
        schedule,
        [
            RAPlan("RA1", "RA1", 1, block_key="block_1"),
            RAPlan("RA2", "RA2", 7, block_key="block_2"),
        ],
        [
            BlockPlan("block_1", "Bloc 1", {0: 60, 1: 0, 2: 0, 3: 0, 4: 0}),
            BlockPlan("block_2", "Bloc 2", {0: 60, 1: 0, 2: 120, 3: 0, 4: 0}),
        ],
    )
    assert rows[0]["ra_hours"] == {"RA1": 0, "RA2": 2}
    assert rows[1]["ra_hours"] == {"RA1": 1, "RA2": 1}
    assert rows[2]["ra_hours"] == {"RA1": 0, "RA2": 2}
    assert rows[3]["ra_hours"] == {"RA1": 0, "RA2": 2}


def test_parallel_blocks_can_transfer_unused_hours_in_reverse_direction():
    """If block 2 finishes early, its spare hours should move to block 1."""

    schedule = build_schedule(
        date(2026, 9, 7),
        date(2026, 9, 21),
        {0: 2, 2: 2},
        set(),
    )
    rows = allocate_ra_hours_by_blocks(
        schedule,
        [
            RAPlan("RA1", "RA1", 7, block_key="block_1"),
            RAPlan("RA2", "RA2", 1, block_key="block_2"),
        ],
        [
            BlockPlan("block_1", "Bloc 1", {0: 60, 1: 0, 2: 0, 3: 0, 4: 0}),
            BlockPlan("block_2", "Bloc 2", {0: 60, 1: 0, 2: 120, 3: 0, 4: 0}),
        ],
    )
    assert rows[0]["ra_hours"] == {"RA1": 1, "RA2": 1}
    assert rows[1]["ra_hours"] == {"RA1": 2, "RA2": 0}
    assert rows[2]["ra_hours"] == {"RA1": 2, "RA2": 0}
    assert rows[3]["ra_hours"] == {"RA1": 2, "RA2": 0}


def test_export_uses_blank_cells_for_zero_ra_hours():
    """Zero-hour RA cells should stay visually empty in the spreadsheet."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4, "RA2": 0},
            }
        ],
        [RAPlan("RA1", "RA1", 4), RAPlan("RA2", "RA2", 0)],
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    merged_ranges = [str(range_ref) for range_ref in sheet.merged_cells.ranges]

    assert sheet.freeze_panes == "D7"
    assert "D1:F1" in merged_ranges
    assert "A2:C2" in merged_ranges
    assert "A6:B6" in merged_ranges
    assert "A8:C8" in merged_ranges
    assert sheet["D1"].value == "MODIFIQUEU LES HORES D'UNA DATA CONCRETA PER\nRECALCULAR AUTOMÀTICAMENT LA RÀTIO D’ACOMPLIMENT"
    assert sheet["D1"].alignment.horizontal == "center"
    assert sheet["D1"].font.bold is True
    assert sheet["D1"].font.color.rgb == "00FFFFFF"
    assert sheet["D1"].fill.fgColor.rgb == "00000000"
    assert sheet.row_dimensions[1].height >= 36
    assert sheet["A2"].value == "Cicle formatiu:"
    assert sheet["D2"].value is None
    assert sheet["A6"].value == "Data"
    assert sheet["A6"].alignment.horizontal == "left"
    assert sheet["A7"].value == "dt."
    assert sheet["A7"].alignment.horizontal == "center"
    assert sheet["B7"].value.date().isoformat() == "2026-09-01"
    assert sheet["B7"].number_format == "DD/MM/YYYY"
    assert sheet["B7"].alignment.horizontal == "center"
    assert sheet["C6"].value == "Hores"
    assert sheet.column_dimensions["C"].width < 16
    assert sheet["C6"].alignment.horizontal == "center"
    assert sheet["C7"].alignment.horizontal == "center"
    assert sheet["D6"].value == "RA1"
    assert sheet["E6"].value == "RA2"
    assert sheet["F6"].value == "Comentaris"
    assert sheet["F6"].alignment.horizontal == "center"
    assert sheet.column_dimensions["F"].width == 42
    assert sheet["D6"].alignment.horizontal == "center"
    assert sheet["D7"].alignment.horizontal == "center"
    assert sheet["F7"].alignment.horizontal == "left"
    assert sheet["D6"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["A6"].border.left.style == "thin"
    assert sheet["D7"].value == 4
    assert sheet["E7"].value is None
    assert sheet["F7"].value is None
    assert sheet["F8"].border.left.style is None
    assert sheet["F9"].border.left.style is None
    assert sheet["F10"].border.left.style is None
    assert sheet["A8"].value == "Hores previstes:"
    assert sheet["A9"].value == "Hores reals:"
    assert sheet["A10"].value == "Acompliment:"
    assert sheet["A8"].border.left.style is None
    assert sheet["C8"].border.right.style is None
    assert sheet["D8"].border.left.style == "thin"
    assert sheet["D8"].value == 4
    assert sheet["E8"].value == 0
    assert sheet["F8"].value == "=SUM(D8:E8)"
    assert sheet["F8"].alignment.horizontal == "center"
    assert sheet["F8"].font.bold is True
    assert sheet["D8"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["E8"].fill.fgColor.rgb == "00D8EBF2"
    assert sheet["D9"].value == "=SUM(D7:D7)"
    assert sheet["E9"].value == "=SUM(E7:E7)"
    assert sheet["F9"].value == "=SUM(D9:E9)"
    assert sheet["F9"].alignment.horizontal == "center"
    assert sheet["D10"].value == '=IFERROR(D9/D8,"")'
    assert sheet["E10"].value == '=IFERROR(E9/E8,"")'
    assert sheet["F10"].value == '=IFERROR(F9/F8,"")'
    assert sheet["F10"].alignment.horizontal == "center"
    assert sheet["D10"].number_format == "0%"
    assert sheet["F10"].number_format == "0%"



def test_export_summary_rows_use_formulas_for_real_and_completion_hours():
    """The XLSX summary should total real hours and compute completion percentages."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 3,
                "ra_hours": {"RA1": 1, "RA2": 2},
            },
            {
                "date": date(2026, 9, 2),
                "weekday": "Dimecres",
                "total_hours": 2,
                "ra_hours": {"RA1": 2, "RA2": 0},
            },
        ],
        [RAPlan("RA1", "RA1", 3), RAPlan("RA2", "RA2", 2)],
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["D9"].value == 3
    assert sheet["E9"].value == 2
    assert sheet["F9"].value == "=SUM(D9:E9)"
    assert sheet["D10"].value == "=SUM(D7:D8)"
    assert sheet["E10"].value == "=SUM(E7:E8)"
    assert sheet["F10"].value == "=SUM(D10:E10)"
    assert sheet["D11"].value == '=IFERROR(D10/D9,"")'
    assert sheet["E11"].value == '=IFERROR(E10/E9,"")'
    assert sheet["F11"].value == '=IFERROR(F10/F9,"")'
    assert sheet["E11"].number_format == "0%"
    assert sheet["F11"].number_format == "0%"



def test_export_marks_summary_formulas_as_ignored_for_adjacent_formula_warnings():
    """The XLSX should suppress Excel's adjacent-formula warning for summary rows."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 3,
                "ra_hours": {"RA1": 1, "RA2": 2},
            },
            {
                "date": date(2026, 9, 2),
                "weekday": "Dimecres",
                "total_hours": 2,
                "ra_hours": {"RA1": 2, "RA2": 0},
            },
        ],
        [RAPlan("RA1", "RA1", 3), RAPlan("RA2", "RA2", 2)],
        {"Camp": "Valor"},
    )

    with ZipFile(workbook_io) as workbook_zip:
        sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert '<ignoredErrors>' in sheet_xml
    assert 'sqref="D10:F10"' in sheet_xml
    assert 'sqref="D11:F11"' in sheet_xml
    assert 'formulaRange="1"' in sheet_xml



def test_export_uses_two_row_ra_headers_when_optional_names_are_present():
    """Optional RA names should expand the calendar header only when needed."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4, "RA2": 0, "RA3": 0},
            }
        ],
        [
            RAPlan("RA1", "Opcional", 4),
            RAPlan("RA2", "RA2", 0),
            RAPlan("RA3", "Opcional", 0),
        ],
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    merged_ranges = [str(range_ref) for range_ref in sheet.merged_cells.ranges]

    assert sheet.freeze_panes == "D8"
    assert "A6:B7" in merged_ranges
    assert "C6:C7" in merged_ranges
    assert "E6:E7" in merged_ranges
    assert "G6:G7" in merged_ranges
    assert sheet["D6"].value == "RA1"
    assert sheet["D6"].border.bottom.style is None
    assert sheet["D7"].value == "Opcional"
    assert sheet["D7"].border.top.style is None
    assert sheet["E6"].value == "RA2"
    assert sheet["E7"].value is None
    assert sheet["F6"].value == "RA3"
    assert sheet["F7"].value == "Opcional"
    assert sheet["G6"].value == "Comentaris"
    assert sheet["A8"].value == "dt."
    assert sheet["B8"].value.date().isoformat() == "2026-09-01"



def test_export_includes_optional_module_fields_above_the_calendar_table():
    """The calendar sheet should reserve editable rows for optional module metadata."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4},
            }
        ],
        [RAPlan("RA1", "RA1", 4)],
        {
            "Cicle formatiu": "Administració i finances",
            "Grup": "-",
            "Codi del mòdul": "MP0440",
            "Nom del mòdul": "-",
        },
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["A2"].value == "Cicle formatiu:"
    assert sheet["D2"].value == "Administració i finances"
    assert sheet["A3"].value == "Grup:"
    assert sheet["D3"].value is None
    assert sheet["A4"].value == "Codi del mòdul:"
    assert sheet["D4"].value == "MP0440"
    assert sheet["A5"].value == "Nom del mòdul:"
    assert sheet["D5"].value is None



def test_export_reuses_palette_colors_after_the_twelfth_ra():
    """RA header colors should wrap around the palette instead of crashing."""

    ras = [RAPlan(f"RA{index}", f"RA{index}", 1) for index in range(1, 15)]
    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 1,
                "ra_hours": {ra.key: (1 if ra.key == "RA1" else 0) for ra in ras},
            }
        ],
        ras,
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["D6"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["O6"].fill.fgColor.rgb == "00EFE0C8"
    assert sheet["P6"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["Q6"].fill.fgColor.rgb == "00D8EBF2"
    assert sheet["R6"].value == "Comentaris"
    assert sheet.column_dimensions["R"].width == 42
    assert sheet["D8"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["Q8"].fill.fgColor.rgb == "00D8EBF2"



def test_export_marks_new_ra_start_rows_with_light_fill():
    """Rows where a new RA starts should be visually highlighted."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 1,
                "ra_hours": {"RA1": 1, "RA2": 0},
            },
            {
                "date": date(2026, 9, 8),
                "weekday": "Dimarts",
                "total_hours": 1,
                "ra_hours": {"RA1": 0, "RA2": 1},
            },
        ],
        [RAPlan("RA1", "RA1", 1), RAPlan("RA2", "RA2", 1)],
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["A7"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["D7"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["E7"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["F7"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["A8"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["D8"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["E8"].fill.fgColor.rgb == "00D8EBF2"
    assert sheet["F8"].fill.fgColor.rgb == "00E9E9E9"



def test_export_parallel_mode_only_marks_rows_where_a_new_ra_first_appears():
    """Parallel planning should not mark rows just because the active mix changes."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 3,
                "ra_hours": {"RA1": 1, "RA2": 2, "RA3": 0},
            },
            {
                "date": date(2026, 9, 2),
                "weekday": "Dimecres",
                "total_hours": 2,
                "ra_hours": {"RA1": 0, "RA2": 2, "RA3": 0},
            },
            {
                "date": date(2026, 9, 9),
                "weekday": "Dimecres",
                "total_hours": 2,
                "ra_hours": {"RA1": 0, "RA2": 1, "RA3": 1},
            },
        ],
        [RAPlan("RA1", "RA1", 1), RAPlan("RA2", "RA2", 5), RAPlan("RA3", "RA3", 1)],
        {"Camp": "Valor"},
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["A7"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["D7"].fill.fgColor.rgb == "00E7EFD8"
    assert sheet["E7"].fill.fgColor.rgb == "00D8EBF2"
    assert sheet["G7"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["A8"].fill.patternType is None
    assert sheet["D8"].fill.patternType is None
    assert sheet["E8"].fill.patternType is None
    assert sheet["G8"].fill.patternType is None
    assert sheet["A9"].fill.fgColor.rgb == "00E9E9E9"
    assert sheet["F9"].fill.fgColor.rgb == "00F7E3D1"
    assert sheet["G9"].fill.fgColor.rgb == "00E9E9E9"


def test_parse_date_input_accepts_dd_mm_yyyy():
    """The web UI should accept dates in the requested Catalan-style format."""

    assert parse_date_input("09/12/2025") == date(2025, 12, 9)


def test_parse_date_input_accepts_single_digit_day_and_month():
    """Manual date entry should not require leading zeroes."""

    assert parse_date_input("7/1/2026") == date(2026, 1, 7)


def test_parse_date_input_accepts_two_digit_year():
    """Manual date entry should also accept a two-digit year."""

    assert parse_date_input("7/11/26") == date(2026, 11, 7)


def test_parse_date_input_keeps_iso_backwards_compatible():
    """Older saved forms or direct POSTs using ISO dates should still work."""

    assert parse_date_input("2025-12-09") == date(2025, 12, 9)


def test_format_display_date_returns_dd_mm_yyyy():
    """Visible dates should always render in DD/MM/YYYY."""

    assert format_display_date(date(2026, 5, 27)) == "27/05/2026"


def test_parse_hour_minute_pair_accepts_twenty_and_thirty_minute_blocks():
    """Parallel block inputs should support minute-level splits in 5-minute steps."""

    assert parse_hour_minute_pair("1", "20") == 80
    assert parse_hour_minute_pair("0", "30") == 30


def test_parse_hour_minute_pair_rejects_non_five_minute_blocks():
    """Minute values outside 5-minute increments should fail validation."""

    with pytest.raises(ValueError, match="5 minuts"):
        parse_hour_minute_pair("1", "32")


def test_format_minutes_label_uses_teacher_friendly_text():
    """Block summaries should be easier to read than decimal hours."""

    assert format_minutes_label(20) == "20 min"
    assert format_minutes_label(80) == "1 h 20 min"


def test_parallel_blocks_support_fractional_daily_capacities():
    """Parallel allocation should honor exact minute capacities per block."""

    schedule = build_schedule(
        date(2026, 9, 6),
        date(2026, 9, 9),
        {0: 2, 1: 2, 2: 0, 3: 0, 4: 0},
        set(),
    )
    rows = allocate_ra_hours_by_blocks(
        schedule,
        [
            RAPlan("RA1", "RA1", 2, block_key="block_1"),
            RAPlan("RA2", "RA2", 2, block_key="block_2"),
        ],
        [
            BlockPlan("block_1", "Bloc 1", {0: 80, 1: 40, 2: 0, 3: 0, 4: 0}),
            BlockPlan("block_2", "Bloc 2", {0: 40, 1: 80, 2: 0, 3: 0, 4: 0}),
        ],
    )

    assert rows[0]["ra_hours"] == {"RA1": 1.33, "RA2": 0.67}
    assert rows[1]["ra_hours"] == {"RA1": 0.67, "RA2": 1.33}


def test_parallel_block_configuration_rejects_a_block_that_exceeds_the_day_total():
    """A single block cannot claim more minutes than the calendar allows that day."""

    with pytest.raises(ValueError, match="superar les hores del calendari"):
        _build_block_plans(
            {
                "block_1_monday_hours": "2",
                "block_1_monday_minutes": "30",
                "block_2_monday_hours": "0",
                "block_2_monday_minutes": "0",
            },
            {0: 2, 1: 0, 2: 0, 3: 0, 4: 0},
            "parallel",
        )


def test_parallel_ra_distribution_accepts_exact_minutes():
    """Parallel-mode RAs should validate against the full minute total, not just whole hours."""

    validate_ra_distribution(
        180,
        [
            RAPlan("RA1", "RA1", 0.67, block_key="block_1", minutes=40),
            RAPlan("RA2", "RA2", 2.33, block_key="block_2", minutes=140),
        ],
        use_minutes=True,
    )


def test_parallel_block_allocation_uses_ra_minutes_exactly():
    """Minute-based RA totals should flow through block allocation without rounding drift."""

    schedule = build_schedule(
        date(2026, 9, 6),
        date(2026, 9, 9),
        {0: 2, 1: 1, 2: 0, 3: 0, 4: 0},
        set(),
    )
    rows = allocate_ra_hours_by_blocks(
        schedule,
        [
            RAPlan("RA1", "RA1", 0.67, block_key="block_1", minutes=40),
            RAPlan("RA2", "RA2", 2.33, block_key="block_2", minutes=140),
        ],
        [
            BlockPlan("block_1", "Bloc 1", {0: 40, 1: 0, 2: 0, 3: 0, 4: 0}),
            BlockPlan("block_2", "Bloc 2", {0: 80, 1: 60, 2: 0, 3: 0, 4: 0}),
        ],
    )

    assert rows[0]["ra_hours"] == {"RA1": 0.67, "RA2": 1.33}
    assert rows[1]["ra_hours"] == {"RA1": 0, "RA2": 1}


def test_week_start_returns_the_monday_of_the_containing_week():
    """Any weekday should resolve to the Monday that starts its calendar week."""

    assert week_start(date(2026, 9, 9)) == date(2026, 9, 7)
    assert week_start(date(2026, 9, 7)) == date(2026, 9, 7)
    assert week_start(date(2026, 9, 13)) == date(2026, 9, 7)


def test_iter_week_starts_covers_every_overlapping_week():
    """The week list should include the Mondays of both the first and last week."""

    assert iter_week_starts(date(2026, 9, 8), date(2026, 9, 21)) == [
        date(2026, 9, 7),
        date(2026, 9, 14),
        date(2026, 9, 21),
    ]


def test_week_has_teaching_potential_is_false_for_a_fully_excluded_week():
    """A week only counts as numbered if at least one weekday survives exclusions."""

    monday = date(2026, 9, 7)
    fully_excluded = {monday + timedelta(days=offset) for offset in range(5)}
    assert week_has_teaching_potential(monday, fully_excluded) is False

    partially_excluded = fully_excluded - {monday}
    assert week_has_teaching_potential(monday, partially_excluded) is True


def test_suggest_week_numbers_skips_fully_excluded_weeks():
    """Vacation weeks should not consume a sequential number."""

    week_starts = [date(2026, 9, 7), date(2026, 9, 14), date(2026, 9, 21)]
    excluded_dates = {date(2026, 9, 14) + timedelta(days=offset) for offset in range(5)}
    suggestions = suggest_week_numbers(week_starts, excluded_dates)
    assert suggestions == {date(2026, 9, 7): 1, date(2026, 9, 21): 2}


def test_week_number_for_date_looks_up_by_week_start():
    """Any date within a numbered week should resolve to that week's number."""

    saved_numbers = {date(2026, 9, 7): 3}
    assert week_number_for_date(date(2026, 9, 9), saved_numbers) == 3
    assert week_number_for_date(date(2026, 9, 21), saved_numbers) is None


def test_is_vacation_week_is_true_for_a_blank_fully_excluded_week():
    """An unnumbered week with zero teaching potential should read as a vacation week."""

    monday = date(2026, 12, 21)
    fully_excluded = {monday + timedelta(days=offset) for offset in range(5)}
    assert is_vacation_week(None, monday, fully_excluded) is True


def test_is_vacation_week_is_false_when_a_number_was_explicitly_assigned():
    """An admin-assigned number always wins, even on an otherwise fully excluded week."""

    monday = date(2026, 12, 21)
    fully_excluded = {monday + timedelta(days=offset) for offset in range(5)}
    assert is_vacation_week(7, monday, fully_excluded) is False


def test_is_vacation_week_is_false_when_the_week_still_has_teaching_potential():
    """A blank week that could still be taught is a real gap, not a vacation week."""

    monday = date(2026, 9, 7)
    assert is_vacation_week(None, monday, set()) is False


def test_format_month_label_uses_catalan_month_names():
    """Month grouping in the admin grid should render in Catalan."""

    assert format_month_label(date(2026, 9, 7)) == "Setembre 2026"


def test_export_adds_leading_week_column_when_requested():
    """Passing include_week_numbers should insert a Set. column before the date."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4},
                "week_number": 5,
            }
        ],
        [RAPlan("RA1", "RA1", 4)],
        {"Camp": "Valor"},
        include_week_numbers=True,
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    merged_ranges = [str(range_ref) for range_ref in sheet.merged_cells.ranges]

    assert sheet.freeze_panes == "E7"
    assert "B6:C6" in merged_ranges
    assert sheet["A6"].value == "Set."
    assert sheet["B6"].value == "Data"
    assert sheet["D6"].value == "Hores"
    assert sheet["E6"].value == "RA1"
    assert sheet["F6"].value == "Comentaris"
    assert sheet["A7"].value == 5
    assert sheet["B7"].value == "dt."
    assert sheet["C7"].value.date().isoformat() == "2026-09-01"
    assert sheet["E7"].value == 4


def test_export_leaves_week_cell_blank_when_that_week_has_no_saved_number():
    """A row with no assigned week number should render a blank cell, not zero."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4},
                "week_number": None,
            }
        ],
        [RAPlan("RA1", "RA1", 4)],
        {"Camp": "Valor"},
        include_week_numbers=True,
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    assert sheet["A7"].value is None


def test_export_week_column_merges_across_both_header_rows_with_optional_names():
    """The Set. header should span both header rows, like the other fixed columns."""

    workbook_io = build_workbook(
        [
            {
                "date": date(2026, 9, 1),
                "weekday": "Dimarts",
                "total_hours": 4,
                "ra_hours": {"RA1": 4},
                "week_number": 1,
            }
        ],
        [RAPlan("RA1", "Opcional", 4)],
        {"Camp": "Valor"},
        include_week_numbers=True,
    )
    workbook = load_workbook(workbook_io)
    sheet = workbook["Calendari"]
    merged_ranges = [str(range_ref) for range_ref in sheet.merged_cells.ranges]

    assert "A6:A7" in merged_ranges
    assert "B6:C7" in merged_ranges
    assert sheet["A6"].value == "Set."
    assert sheet["A8"].value == 1
